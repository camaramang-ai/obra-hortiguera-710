import openpyxl
import os

base = '/Users/PabloMan/Hermes/projects/active/obra-hortiguera-710/presupuestos'

files = [
    'PRESUPUESTO 1 maxima ESTIM PAGOS HORTIGUERA 710 1ER PISO     13 05 26.xlsx',
    'PRESUPUESTO 2 minima ESTIM PAGOS 1HORTIGUERA 710 1ER PISO     26 05 26.xlsx'
]

for fname in files:
    path = os.path.join(base, fname)
    print(f'\n{"="*70}')
    print(f'📄 {fname}')
    print(f'{"="*70}')
    
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f'Sheets: {wb.sheetnames}')
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f'\n--- Sheet: {sname} ---')
        print(f'Dimensions: {ws.dimensions}')
        print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            vals = [cell.value for cell in row]
            # Skip completely empty rows
            if all(v is None for v in vals):
                continue
            # Show row number and values
            row_num = row[0].row
            # Format: show non-None values with their column letter
            formatted = []
            for cell in row:
                if cell.value is not None:
                    formatted.append(f'{cell.column_letter}:{cell.value}')
            print(f'  R{row_num}: {" | ".join(formatted)}')
    
    wb.close()
