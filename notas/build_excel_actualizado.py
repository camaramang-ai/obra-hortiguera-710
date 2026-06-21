#!/usr/bin/env python3
"""
Excel visual: Obra Hortiguera 710 — Comparativa + Proveedores + Contactos + Resumen
Optimizado para lectura rápida: colores, secciones, totales, estado visual.
"""
import openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()

# ── PALETA DE COLORES ──────────────────────────────────────────────────
NAVY    = "1F4E79"
BLUE    = "2E75B6"
LBLUE   = "D6E4F0"
GREEN   = "C6EFCE"
DGREEN  = "006100"
RED     = "FFC7CE"
DRED    = "9C0006"
GOLD    = "FFD700"
AMBER   = "FFF2CC"
ORANGE  = "FCE4D6"
WHITE   = "FFFFFF"
LGRAY   = "F2F2F2"
MGRAY   = "D9D9D9"
TEAL    = "B4D6D0"
PURPLE  = "D9D2E9"

HDR_FILL   = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SUBH_FILL  = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
HDR_FONT   = Font(bold=True, color=WHITE, size=10, name="Calibri")
SUBH_FONT  = Font(bold=True, color=WHITE, size=9, name="Calibri")
BOLD       = Font(bold=True, size=10, name="Calibri")
BOLD_BIG   = Font(bold=True, size=14, name="Calibri", color=NAVY)
ITALIC     = Font(italic=True, size=9, color="555555", name="Calibri")
BEST_FILL  = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
WARN_FILL  = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
AMBER_FILL = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")
ORANGE_FILL= PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
TEAL_FILL  = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
PURPLE_FILL= PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
GOLD_FILL  = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
GRAY_FILL  = PatternFill(start_color=LGRAY, end_color=LGRAY, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
THIN    = Border(left=Side(style='thin', color=MGRAY),
                 right=Side(style='thin', color=MGRAY),
                 top=Side(style='thin', color=MGRAY),
                 bottom=Side(style='thin', color=MGRAY))
BOTTOM_BOLD = Border(bottom=Side(style='medium', color=NAVY))
NO_BORDER = Border()

def hdr_cell(ws, row, col, text, w=12):
    c = ws.cell(row=row, column=col, value=text)
    c.font = SUBH_FONT if row > 3 else HDR_FONT
    c.fill = SUBH_FILL if row > 3 else HDR_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = THIN
    return c

def body_cell(ws, row, col, val, bold=False, fill=None, align='center', fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=9, name="Calibri", color=DGREEN if bold else "333333")
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=(align=='left'))
    c.border = THIN
    if fill:
        c.fill = fill
    elif row % 2 == 0:
        c.fill = GRAY_FILL
    if fmt:
        c.number_format = fmt
    return c

def section_row(ws, row, cols, text, fill=None):
    """Merge row for a section header."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    c.fill = fill or SUBH_FILL
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = THIN
    for i in range(2, cols+1):
        ws.cell(row=row, column=i).border = THIN
        ws.cell(row=row, column=i).fill = fill or SUBH_FILL

# ══════════════════════════════════════════════════════════════════════
# SHEET 1: COMPARATIVA
# ══════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = " Comparativa "
ws.sheet_properties.tabColor = NAVY

COLS = 11
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COLS)
c = ws.cell(row=1, column=1, value="OBRA HORTIGUERA 710 · COMPARATIVA COMPLETA")
c.font = BOLD_BIG
c.alignment = Alignment(horizontal='left', vertical='center')

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COLS)
c = ws.cell(row=2, column=1, value="Precios c/IVA incluido  |  Barugel = referencia  |  33 productos de 4 hojas  |  7 proveedores contactados")
c.font = ITALIC

hdrs = ["Item","Producto","Cant.","Hoja","Barugel c/IVA","Mejor Precio","Ahorro","%","Mejor Proveedor","Contactado","Obs"]
for i, h in enumerate(hdrs, 1):
    hdr_cell(ws, 3, i, h)

ws.row_dimensions[3].height = 30

# DATA: [item, prod, cant, hoja, barugel, mejor, ahorro_s, ahorro_pct, proveedor, status, obs]
# status: ENVIADO, PENDIENTE, OK, IRRELEVANTE
DATA = [
    # Hoja 1 - Revestimientos
    None,  # section marker
    ["D.1","Portobello Art Martase (m²)","1.44","H1","$57,819","$57,819","$0","0%","Barugel","ENVIADO","No encontrado en otros proveedores"],
    ["D.2","Tendenza Black Cement (m²)","4.95","H1","$22,442","$21,177","-$1,265","-6%","Ferrocons","ENVIADO","ML #1 a $21.177"],
    ["D.3","Eliane Forma Branco 32x60 (m²)","30.10","H1","$19,525","$19,525","$0","0%","Barugel / Banchero","ENVIADO","Banchero distribuye Eliane"],
    ["D.4","Portobello Nord Ris Ripado 30x90 (m²)","4.02","H1","$57,536","$47,337","-$10,199","-18%","Ferrocons","ENVIADO","—"],
    ["D.5","Vite Liscio Light Eco Grey 20x120 (m²)","4.32","H1","$43,406","$28,872","—","—","ML (formato dif.)","ENVIADO","120x120 vs 60x120"],
    ["D.6","Pinto Vivant Luge Suave 7x24 (m²)","2.04","H1","$117,078","$60,984","-$56,094","-48%","Edificor ⭐","ENVIADO","¡Casi 50% más barato!"],
    ["D.7","Pinto PBG Vivant Suage 7x24 (m²)","3.06","H1","$117,078","$60,984","-$56,094","-48%","Edificor ⭐","ENVIADO","—"],
    ["K.1","Klaukol Adhesivo Porcelanato x25kg","26 un","H1","$33,489","$18,200","-$15,289","-46%","Bercomat ⭐","ENVIADO","Mejor precio encontrado"],
    None,
    # Hoja 2 - Pastinas
    ["P.1","Klaukol Pastina Fluida Blanca 5kg","1","H2","$62,471","$12,872","-$49,599","-79%","ML ⭐","ENVIADO","Muy por debajo de Barugel"],
    ["P.2","Klaukol Pastina Fluida Gris Claro 5kg","1","H2","$31,236","$14,000","-$17,236","-55%","ML","ENVIADO","—"],
    ["P.3","Klaukol Pastina AP Perla 5kg","2","H2","$33,229","$30,888","-$2,341","-7%","La Clarita","ENVIADO","—"],
    None,
    # Hoja 2 - Sanitarios Ferrum Bari
    ["S.1","Inodoro Largo Ferrum Bari Blanco","1","H2","$207,024","$163,695","-$43,329","-21%","El Amigo ⭐","ENVIADO","Consultado a 7 proveedores"],
    ["S.2","Depósito Ferrum Bari Dual Apoyar","1","H2","$199,312","$142,194","-$57,118","-29%","El Amigo ⭐","ENVIADO","—"],
    ["S.3","Bidet Ferrum Bari 3 Agujeros","1","H2","$152,316","$103,248","-$49,068","-32%","ML ⭐","ENVIADO","—"],
    ["S.4","Tapa Asiento Ferrum Bari Madera","1","H2","$70,438","$32,912","-$37,526","-53%","Merlino ⭐","ENVIADO","—"],
    None,
    # Hoja 2 - Vanitorys
    ["V.1","Schneider Vanitory City 80cm Colg","1","H2","$676,436","$599,000","-$77,436","-11%","ML","ENVIADO","Pendiente verificar Dimora"],
    ["V.2","Schneider Vanitory Rivo 60cm Colg","1","H2","$534,380","$307,703","-$226,677","-42%","Merlino ⭐","ENVIADO","—"],
    None,
    # Hoja 2 - Bañera/Desague
    ["B.1","FV Bañera Enoxado 150x70","1","H2","$242,897","$242,897","$0","0%","Barugel","ENVIADO","Mejor precio en Barugel"],
    ["B.2","FV Desagüe Lineal 60cm","1","H2","$227,292","$190,642","-$36,650","-16%","Gili y Cía","ENVIADO","—"],
    None,
    # Hoja 3 - Grifería FV Mauna
    ["G.1","FV Lavatorio Mauna Cromo (JG)","3 JG","H3","$198,599","$185,897","-$12,702","-6%","NIMAT","ENVIADO","Banchero, Resta, Blaisten consultados"],
    ["G.2","FV Bidet Mauna Cromo (JG)","2 JG","H3","$216,636","$202,781","-$13,855","-6%","NIMAT","ENVIADO","—"],
    ["G.3","FV Ducha Mauna c/Transfer. (JG)","2 JG","H3","$226,815","$144,690","-$82,125","-36%","NIMAT ⭐","ENVIADO","Ducha mucho más barata en NIMAT"],
    ["G.4","FV Duchamatic c/Brazo Cromo","2 un","H3","$274,764","$215,511","-$59,253","-22%","Resta","ENVIADO","-5% transferencia"],
    None,
    # Hoja 3 - Accesorios FV
    ["A.1","FV Portarrollos Cromo C/0167","3","H3","$43,853","$17,700","-$26,153","-60%","ML Arizona ⭐","ENVIADO","—"],
    ["A.2","FV Toallero Barra Cromo C/0405","3","H3","$73,089","$30,000","-$43,089","-59%","ML Arizona ⭐","ENVIADO","—"],
    ["A.3","FV Percha Cromo C/0166","5","H3","$27,289","$12,000","-$15,289","-56%","ML Arizona ⭐","ENVIADO","—"],
    None,
    # Hoja 4 - Espejos
    ["H4.1","Espejo Reflejar Rect. 100x80 4mm","1","H4","$297,814","$259,100","-$38,714","-13%","ML","ENVIADO","Banchero consultado"],
    ["H4.2","Espejo 20x30 Oxboro Ilumin 220V","1","H4","$282,385","$200,000","-$82,385","-29%","ML / Schneider","ENVIADO","—"],
    None,
    # Hoja 4 - Roca GAP
    ["H4.3","Indoor Largo GAP Blanco Roca","1","H4","$500,148","$500,148","$0","0%","Barugel","ENVIADO","Banchero distribuye Roca"],
    ["H4.4","Alimentación Efectiva 3/6 L","1","H4","$453,198","$453,198","$0","0%","Barugel","ENVIADO","—"],
    ["H4.5","Bidet 3 Ag. Roca La Gama","1","H4","$169,411","$169,411","$0","0%","Barugel","ENVIADO","—"],
    ["H4.6","Tapa Amortiguador Slim","1","H4","$62,355","$39,270","-$23,085","-37%","Remateriales ⭐","ENVIADO","—"],
    ["H4.7","Portobello Zesty Light 45x120 Text","6.44 m²","H4","—","—","—","—","Consultar","ENVIADO","Incluido en consulta Banchero"],
]

SECTION_LABELS = {
    1: "HOJA 1 — Revestimientos y Porcelanatos",
    2: "HOJA 2 — Pastinas",
    3: "HOJA 2 — Sanitarios Ferrum Bari",
    4: "HOJA 2 — Vanitorys Schneider",
    5: "HOJA 2 — Bañera y Desagüe FV",
    6: "HOJA 3 — Grifería FV Mauna",
    7: "HOJA 3 — Accesorios FV",
    8: "HOJA 4 — Espejos",
    9: "HOJA 4 — Roca GAP y Zesty",
}
SECTION_FILLS = {
    1: PatternFill(start_color=LBLUE, end_color=LBLUE, fill_type="solid"),
    2: PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid"),
    3: PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid"),
    4: PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid"),
    5: PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid"),
    6: PatternFill(start_color=LBLUE, end_color=LBLUE, fill_type="solid"),
    7: PatternFill(start_color=LBLUE, end_color=LBLUE, fill_type="solid"),
    8: PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid"),
    9: PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid"),
}

sec_idx = 0
r = 4
for item in DATA:
    if item is None:
        sec_idx += 1
        continue  # section markers placed before data
    # Check if previous was None (new section)
    # We'll use a different approach: track sections via indices

# Reset and write with proper sections
sec_idx = 0
r = 4
for i, item in enumerate(DATA):
    if item is None:
        sec_idx += 1
        label = SECTION_LABELS.get(sec_idx, "")
        fill = SECTION_FILLS.get(sec_idx, SUBH_FILL)
        section_row(ws, r, COLS, f"  {label}", fill)
        ws.row_dimensions[r].height = 24
        r += 1
        continue

    code, prod, cant, hoja, barugel, mejor, ahorro_s, ahorro_p, prov, status, obs = item
    row_data = [code, prod, cant, hoja, barugel, mejor, ahorro_s, ahorro_p, prov, status, obs]

    # Determine status color
    status_fill = None
    if status == "ENVIADO":
        status_fill = AMBER_FILL
    elif status == "OK" or status == "⭐":
        status_fill = BEST_FILL
    elif status == "IRRELEVANTE":
        status_fill = GRAY_FILL

    # Determine if ahorro is significant
    ahorro_fill = None
    try:
        pct_str = ahorro_p.replace('%','').replace('—','0').strip()
        if pct_str and float(pct_str) < 0:
            ahorro_fill = BEST_FILL
    except:
        pass

    for c, val in enumerate(row_data, 1):
        align = 'left' if c in (2, 9, 11) else 'center'
        is_ahorro_col = c in (7, 8)
        fill = ahorro_fill if (is_ahorro_col and ahorro_fill) else status_fill if c == 10 else None
        is_star = c == 9 and '⭐' in str(val)
        body_cell(ws, r, c, val, bold=is_star, fill=fill, align=align)

    ws.row_dimensions[r].height = 22
    r += 1

# Freeze panes
ws.freeze_panes = 'A4'

# Auto-width
col_widths = {1:7, 2:42, 3:10, 4:6, 5:13, 6:13, 7:14, 8:8, 9:22, 10:12, 11:38}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ══════════════════════════════════════════════════════════════════════
# SHEET 2: PROVEEDORES
# ══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet(" Proveedores ")
ws2.sheet_properties.tabColor = BLUE

ws2.merge_cells('A1:G1')
c = ws2.cell(row=1, column=1, value="PROVEEDORES RELEVADOS · 16 proveedores")
c.font = BOLD_BIG

ws2.merge_cells('A2:G2')
c = ws2.cell(row=2, column=1, value="Verde = ya presupuestado  |  Ámbar = consulta enviada  |  Naranja = pendiente")
c.font = ITALIC

phdrs = ["#","Proveedor","Contacto","Canal","Rubro","Ubicación / Distancia","Estado"]
for i, h in enumerate(phdrs, 1):
    hdr_cell(ws2, 3, i, h)
ws2.row_dimensions[3].height = 28

PROVS = [
    ["1","Barugel Azulay & Cía","Marcos Badillo\nmbadillo@barugel.com.ar","Email / WA","Todo","Alberdi 3701, Floresta (~2km)","✅ Presupuesto base"],
    ["2","Blaisten","ventasweb@blaisten.com","Email","Revest., sanit., grifería","Alberdi 3928, Floresta (~2km)","📤 Consulta enviada"],
    ["3","Bercomat","consultas@bercomat.com","Email","Sanit., grifería, adhesivos","Varias GBA","📤 Consulta enviada"],
    ["4","Easy","ventasecommerce@cencosud.com.ar","Email","Todo construcción","Palermo, CABA","📤 Consulta enviada"],
    ["5","Resta Sanitarios","info@sanitariosresta.com.ar","Email","FV, Roca, sanitarios","CABA","📤 Consulta enviada"],
    ["6","Banchero Sanitarios ⭐","ventas@bancherosanitarios.com.ar","Email","FV, Ferrum, Roca, Eliane","Ángel Gallardo 866, Caballito","📤 Consulta enviada"],
    ["7","Caballito Sanitarios","info@caballitosanitarios.com.ar","Email / Pres.","Ferrum, FV, sanitarios","Donato Álvarez 64 (~5 cdas)","📤 Consulta enviada"],
    ["8","Generación Sanitaria ⭐","gsrosario@yahoo.com","Email","Ferrum, sanitarios","Rosario 555 (~12 cdas)","📤 Consulta enviada"],
    ["9","Corralón Caballito","11 3301-1402","WhatsApp","Mat. generales, adhesivos","Donato Álvarez 174 (~5 cdas)","⏳ Pendiente WA"],
    ["10","Oeste Sanitarios","bit.ly/OesteSanitario-Tienda","WhatsApp","Sanit., grifería, repuestos","Colpayo 412 (~8 cdas)","⏳ Pendiente WA"],
    ["11","Corralón Premium II","4903-9666","Presencial","Mat. construcción","Av. J. M. Moreno 487 (~10 cdas)","⏳ Pendiente"],
    ["12","NIMAT","ventas@nimat.com.ar","Email","Grifería FV Mauna","CABA","✅ Mejor ducha Mauna"],
    ["13","El Amigo","ventas@elamigo.com.ar","Email","Sanitarios Ferrum","CABA","✅ Mejor Ferrum Bari"],
    ["14","Merlino","ventas@merlino.com.ar","Email","Vanitory, muebles baño","CABA","✅ Mejor V. Rivo"],
    ["15","Edificor","ventas@edificor.com.ar","Email","Revest. porcelánicos","CABA","✅ Mejor Pinto Vivant"],
    ["16","Ferrocons","ventas@ferrocons.com.ar","Email","Revestimientos","CABA","✅ Buen precio Nord Ris"],
]

STATUS_COLORS = {
    "✅": BEST_FILL,
    "📤": AMBER_FILL,
    "⏳": ORANGE_FILL,
}

for r_idx, row in enumerate(PROVS, 4):
    for c, val in enumerate(row, 1):
        align = 'left' if c in (2,3,5,6,7) else 'center'
        fill = None
        if c == 7:
            for prefix, pf in STATUS_COLORS.items():
                if str(val).startswith(prefix):
                    fill = pf
                    break
        body_cell(ws2, r_idx, c, val, align=align, fill=fill)
    ws2.row_dimensions[r_idx].height = 30 if '\n' in str(row[2]) else 22

ws2.freeze_panes = 'A4'
col_w2 = {1:5, 2:24, 3:32, 4:14, 5:30, 6:30, 7:22}
for col, w in col_w2.items():
    ws2.column_dimensions[get_column_letter(col)].width = w

# ══════════════════════════════════════════════════════════════════════
# SHEET 3: WHATSAPP
# ══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet(" WhatsApp ")
ws3.sheet_properties.tabColor = GREEN

ws3.merge_cells('A1:E1')
c = ws3.cell(row=1, column=1, value="CONTACTOS WHATSAPP · Copiar y enviar")
c.font = BOLD_BIG

ws3.merge_cells('A2:E2')
c = ws3.cell(row=2, column=1, value="Pendientes de contacto. Los textos están listos para copiar y pegar.")
c.font = ITALIC

whdrs = ["Proveedor","Dirección","Distancia","Contacto","Texto para copiar"]
for i, h in enumerate(whdrs, 1):
    hdr_cell(ws3, 3, i, h)

WCONTACTS = [
    ["Corralón Caballito","Donato Álvarez 174","~5 cdas","WA: 11 3301-1402",
     "Hola, cómo están? Soy Pablo, estoy haciendo una remodelación en Caballito. Quería consultarles si tienen: Klaukol Adhesivo Fluido Porcelanato x25kg (26 un), Klaukol Pastina Fluida Blanca 5kg (1), Gris Claro 5kg (1), Perla 5kg (2). Si no tienen alguno, opciones equivalentes. Precio, stock, envío. Gracias!"],
    ["Oeste Sanitarios","Colpayo 412","~8 cdas","bit.ly/OesteSanitario-Tienda",
     "Hola, cómo están? Soy Pablo, estoy con una obra en Caballito. Quería presupuesto de sanitarios Ferrum Bari (inodoro, depósito, bidet, tapa) y grifería FV Mauna (lavatorio, bidet, ducha). Si no tienen exacto, opciones equivalentes. Precio, stock, flete, descuento efectivo. Gracias!"],
    ["Corralón Premium II","Av. J. M. Moreno 487","~10 cdas","4903-9666 (tel)",
     "Ir en persona. Preguntar por adhesivo Klaukol y pastinas."],
]

for r_idx, row in enumerate(WCONTACTS, 4):
    for c, val in enumerate(row, 1):
        align = 'left'
        body_cell(ws3, r_idx, c, val, align=align)
    ws3.row_dimensions[r_idx].height = 60 if r_idx <= 5 else 30

ws3.freeze_panes = 'A4'
ws3.column_dimensions['A'].width = 24
ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 28
ws3.column_dimensions['E'].width = 85

# ══════════════════════════════════════════════════════════════════════
# SHEET 4: RESUMEN
# ══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet(" Resumen ")
ws4.sheet_properties.tabColor = GOLD

ws4.merge_cells('A1:G1')
c = ws4.cell(row=1, column=1, value="RESUMEN DE AHORRO · Gestión de compras")
c.font = BOLD_BIG

ws4.merge_cells('A2:G2')
c = ws4.cell(row=2, column=1, value="Ahorro potencial comprando al mejor precio por producto. No incluye flete.")
c.font = ITALIC

rhdrs = ["Categoría","Costo Barugel","Mejor Precio","Ahorro","%","Ahorro Anualizado","Estado"]
for i, h in enumerate(rhdrs, 1):
    hdr_cell(ws4, 3, i, h)
ws4.row_dimensions[3].height = 28

RESUMEN = [
    ["Revestimientos (sin Cerdisa)","$2,700,156","$1,898,000","$802,000","30%","—","📤 4 proveedores consultados"],
    ["Pastinas Klaukol","$126,937","$58,000","$69,000","54%","—","📤 Email + WA pendiente"],
    ["Set Ferrum Bari (4 pzas)","$629,090","$442,049","$187,000","30%","—","📤 7 proveedores consultados"],
    ["Vanitory Schneider (2 un)","$1,210,816","$765,357","$445,000","37%","—","📤 Incluido en consultas"],
    ["Bañera + Desagüe FV","$470,189","$433,539","$37,000","8%","—","📤 Incluido"],
    ["Grifería FV Mauna (3 JG)","$642,050","$533,368","$109,000","17%","—","📤 Banchero, Resta, Blaisten"],
    ["Duchamatic FV (2 un)","$549,528","$431,022","$119,000","22%","—","📤 Resta consultado"],
    ["Accesorios FV (3 items)","$144,231","$60,000","$84,000","58% ⭐","—","📤 Incluido"],
    ["Espejos (2 un)","$580,199","$459,100","$121,000","21%","—","📤 Banchero consultado"],
    ["Set Roca GAP + Zesty","$1,185,112","$1,162,000","$23,000","2%","—","📤 Banchero (distrib. Roca)"],
]

total_bar = 0
total_mejor = 0
for r_idx, row in enumerate(RESUMEN, 4):
    for c, val in enumerate(row, 1):
        align = 'center'
        is_savings = c in (4, 5)
        fill = None
        if is_savings:
            try:
                pct = float(str(val).replace('%','').replace('⭐','').strip())
                if pct >= 20:
                    fill = BEST_FILL
            except:
                pass
        if c == 4 and str(val) != "$0" and str(val) != "—":
            fill = BEST_FILL if not fill else fill
        body_cell(ws4, r_idx, c, val, align=align, fill=fill)
    # Sum totals
    try:
        v = row[1].replace('$','').replace(',','')
        total_bar += int(v)
    except: pass
    try:
        v = row[2].replace('$','').replace(',','')
        total_mejor += int(v)
    except: pass

# Total row
tr = 4 + len(RESUMEN) + 1
total_vals = ["TOTAL GENERAL", f"${total_bar:,}", f"${total_mejor:,}",
              f"${total_bar - total_mejor:,}", f"{int((1-total_mejor/total_bar)*100)}%",
              "—", f"Ahorro: ${total_bar - total_mejor:,}"]
for c, val in enumerate(total_vals, 1):
    cell = ws4.cell(row=tr, column=c, value=val)
    cell.font = Font(bold=True, size=11, name="Calibri")
    cell.fill = GOLD_FILL
    cell.border = THIN
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Info box
tr += 2
info_items = [
    "🔷 Ítem D Arquitectos (límite): $8,772,420  |  Barugel neto total: $8,092,309  |  Margen libre: $680,111 ✅",
    "ℹ️  Cerdisa Deep Sun: precio ya bueno en Barugel. Verificar stock con Marcos Badillo.",
    "📤 Proveedores contactados vía email (7): Blaisten, Bercomat, Easy, Resta, Banchero, Caballito Sanitarios, Generación Sanitaria",
    "⏳ Pendientes WhatsApp (2): Corralón Caballito (11 3301-1402), Oeste Sanitarios (bit.ly/OesteSanitario-Tienda)",
    "⏳ Pendiente presencial (1): Corralón Premium II (Av. J. M. Moreno 487)",
]
for i, txt in enumerate(info_items):
    r_info = tr + i
    ws4.merge_cells(start_row=r_info, start_column=1, end_row=r_info, end_column=7)
    c = ws4.cell(row=r_info, column=1, value=txt)
    c.font = Font(size=9, name="Calibri", color=NAVY if i==0 else "444444")
    c.alignment = Alignment(horizontal='left', vertical='center')

ws4.freeze_panes = 'A4'
col_w4 = {1:30, 2:16, 3:16, 4:16, 5:10, 6:16, 7:34}
for col, w in col_w4.items():
    ws4.column_dimensions[get_column_letter(col)].width = w

# ═══════════════════════════════════════════════
# SAVE + UPLOAD TO DRIVE
# ═══════════════════════════════════════════════
OUT = os.path.expanduser("~/Hermes/projects/active/obra-hortiguera-710/presupuestos/comparativa_completa.xlsx")
wb.save(OUT)
print(f"✅ Excel guardado: {OUT}")
print(f"   Pestañas: {wb.sheetnames}")

# ── Upload to Google Drive ──
try:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    TK = os.path.expanduser("~/.hermes/google_token.json")
    DSCOPES = ["https://www.googleapis.com/auth/drive", "https://www.googleapis.com/auth/drive.file"]
    creds = Credentials.from_authorized_user_file(TK, DSCOPES)
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    drive = build("drive", "v3", credentials=creds)

    FOLDER_ID = "1huHlteCLIUJ4bASFNOQUbnzxhn3aVSuq"
    FILE_NAME = "comparativa_completa.xlsx"
    query = f"name='{FILE_NAME}' and '{FOLDER_ID}' in parents and trashed=false"
    existing = drive.files().list(q=query, spaces='drive', fields='files(id)').execute()
    files = existing.get('files', [])
    media = MediaFileUpload(OUT, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
    if files:
        drive.files().update(fileId=files[0]['id'], media_body=media).execute()
        print("   ✅ Drive: actualizado")
    else:
        drive.files().create(body={'name': FILE_NAME, 'parents': [FOLDER_ID]}, media_body=media).execute()
        print("   ✅ Drive: subido")
except Exception as e:
    print(f"   ⚠️  No se pudo subir a Drive: {e}")
