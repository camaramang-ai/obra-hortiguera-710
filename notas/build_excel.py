#!/usr/bin/env python3
"""
Build Excel comparativo: Obra Hortiguera 710
Compara precios de los mismos items de Barugel contra 6 proveedores de CABA.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()

# ─── Colors & Styles ───────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BEST_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data_cell(cell, is_best=False, is_total=False):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    if is_best:
        cell.fill = BEST_FILL
        cell.font = Font(bold=True)
    elif is_total:
        cell.font = Font(bold=True, size=11)

def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    max_len = max(max_len, len(line))
        adjusted = min(max_len + 3, 45)
        ws.column_dimensions[col_letter].width = max(adjusted, 10)

# ═══════════════════════════════════════════════════════════════
# SHEET 1: PORTADA / REFERENCIAS
# ═══════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "Portada"
ws0.merge_cells('A1:F1')
ws0.cell(row=1, column=1, value="OBRA HORTIGUERA 710 — COMPARATIVA DE PRECIOS").font = Font(bold=True, size=16)
ws0.merge_cells('A2:F2')
ws0.cell(row=2, column=1, value="Remodelación 1er Piso — Caballito, CABA — Junio 2026").font = Font(size=12, italic=True)

ws0.cell(row=4, column=1, value="REFERENCIAS PRESUPUESTARIAS").font = Font(bold=True, size=13)
headers_ref = ["Concepto", "Presup. Mínima", "Presup. Máxima", "Máxima + Pinotea", "Barugel (Neto)", "Barugel (c/IVA)"]
for i, h in enumerate(headers_ref, 1):
    ws0.cell(row=5, column=i, value=h)
style_header_row(ws0, 5, 6)

data_ref = [
    ["Ítem D (Revest., Artef. y Grifería)", "$3,496,000", "$8,772,420", "$8,772,420", "$8,092,309", "$9,791,694"],
    ["Costo Obra Total", "$49,949,285", "$58,081,705", "$59,521,705", "—", "—"],
    ["Total General (c/honorarios)", "$54,944,214", "$63,889,876", "$65,473,553", "—", "—"],
]
for r, row_data in enumerate(data_ref, 6):
    for c, val in enumerate(row_data, 1):
        cell = ws0.cell(row=r, column=c, value=val)
        style_data_cell(cell, is_total=(r == 8))
        if c == 5 and r == 6:
            cell.fill = BEST_FILL  # Barugel debajo del límite
        if r == 6 and c == 4:
            cell.fill = GOLD_FILL  # Límite máximo

ws0.cell(row=10, column=1, value="LÍMITE DE GASTOS (Ítem D): $8,772,420 — Barugel está $680K por debajo ✅").font = Font(bold=True, color="006100")
ws0.cell(row=11, column=1, value="EXCEPCIÓN: Cerdisa Deep Sun 60x120 (ítem 34616630) — NO comparar, ya está a buen precio en Barugel").font = Font(bold=True, color="C00000")

ws0.cell(row=13, column=1, value="PROVEEDORES RELEVADOS").font = Font(bold=True, size=13)
prov_headers = ["#", "Proveedor", "Dirección / Sucursal", "Especialidad"]
for i, h in enumerate(prov_headers, 1):
    ws0.cell(row=14, column=i, value=h)
style_header_row(ws0, 14, 4)

proveedores = [
    ["1", "Barugel Azulay & Cía", "Av. Alberdi 3701, Floresta (~2km)", "Revestimientos, Ferrum, Schneider (base)"],
    ["2", "Blaisten", "Av. Alberdi 3928, Floresta (~2km)", "Sanitarios Ferrum, Klaukol, vanitory"],
    ["3", "Familia Bercomat", "Varias sucursales GBA", "Grifería FV, adhesivos, materiales"],
    ["4", "Caballito Sanitarios", "Donato Álvarez 64, Caballito (~5 cuadras)", "Ferrum + FV completo"],
    ["5", "Easy", "Av. Int. Francisco Rabanal 4301 / Haedo", "KLaukol, Ferrum, financiación"],
    ["6", "MercadoLibre (varios)", "Online", "Precios de referencia spot"],
]
for r, row_data in enumerate(proveedores, 15):
    for c, val in enumerate(row_data, 1):
        cell = ws0.cell(row=r, column=c, value=val)
        style_data_cell(cell)

ws0.cell(row=22, column=1, value="NOTA: Todos los precios en ARS. Se muestra precio NETO (sin IVA) salvo que se indique lo contrario.").font = Font(italic=True, size=10)
ws0.cell(row=23, column=1, value="IVA = 21%. Precio c/IVA = Neto × 1.21. Factura A siempre que sea posible (IVA es crédito fiscal).").font = Font(italic=True, size=10)

auto_width(ws0)
ws0.sheet_view.showGridLines = False

# ═══════════════════════════════════════════════════════════════
# SHEET 2: HOJA 1 — REVESTIMIENTOS Y PORCELANATOS
# ═══════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Hoja1 - Revestimientos")

# Title
ws1.merge_cells('A1:H1')
ws1.cell(row=1, column=1, value="HOJA 1 — REVESTIMIENTOS Y PORCELANATOS (netos sin IVA)").font = Font(bold=True, size=14)

headers1 = ["Item", "Producto", "Cant.", "Barugel (neto)", "Blaisten", "Bercomat", "Easy", "ML / Otro", "Mejor Precio", "Ahorro vs Barugel"]
for i, h in enumerate(headers1, 1):
    ws1.cell(row=2, column=i, value=h)
style_header_row(ws1, 2, len(headers1))

# Data rows — each: [id, producto, cant, barugel, blaisten, bercomat, easy, ml_otro, mejor, ahorro]
# Prices are NETOS (sin IVA)
revest_data = [
    ["D.1", "Portobello Revest. Art Martase (m²)", "1.44 m²",
     "$47,784", "N/D", "N/D", "N/D", "N/D",
     "$47,784 (Barugel)", "$0"],
    ["D.2", "Porcelanato Tendenza Black Cement (m²)", "4.95 m²",
     "$18,547", "N/D", "N/D", "N/D", "Ferrocons: $18,267",
     "$18,267 (Ferrocons)", "-$280"],
    ["D.3", "Eliane Cerámica Forma Branco 32x60 (m²)", "30.10 m²",
     "$16,136", "N/D", "N/D", "N/D", "Corralón Mdo: $19,922",
     "$16,136 (Barugel)", "$0"],
    ["D.4", "Portobello Nord Ris Decor Ripado 30x90 (m²)", "4.02 m²",
     "$47,550", "N/D", "N/D", "N/D", "Ferrocons: $47,337",
     "$47,337 (Ferrocons)", "-$213"],
    ["D.5", "Vite Porcelanato Liscio Light Eco Grey (m²)*", "4.32 m²",
     "$35,873", "$34,661 (60x120)", "N/D", "N/D", "GAP Haus: $31,642 (s/stk)",
     "$31,642 (GAP, s/stk)", "—"],
    ["D.6", "Pinto Vivant Luge Suave 7x24 (m²)", "2.04 m²",
     "$96,759", "N/D", "N/D", "N/D", "Edificor: $50,400",
     "$50,400 (Edificor)", "-$46,359 ⚡"],
    ["D.7", "Pinto PBG Vivant Suage 7x24 (m²)", "3.06 m²",
     "$96,759", "N/D", "N/D", "N/D", "Edificor: $50,400",
     "$50,400 (Edificor)", "-$46,359 ⚡"],
    ["K.1", "Klaukol Adhesivo Fluido Porcelanato x25kg (un)", "26 un",
     "$27,677", "N/D", "~$18,200†", "$27,900", "Distabile: $22,902",
     "$18,200 (Bercomat)", "-$9,477 c/u ⚡"],
]

for r, row_data in enumerate(revest_data, 3):
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        is_best = (c == 9)  # Mejor precio column
        is_total = (c in [4, 9, 10])
        style_data_cell(cell, is_best=is_best, is_total=is_total)
        # Highlight biggest savings
        if c == 10 and '⚡' in str(val):
            cell.fill = BEST_FILL

# Subtotals row
sub_row = 3 + len(revest_data)
ws1.cell(row=sub_row, column=1, value="SUBTOTAL (neto)").font = Font(bold=True, size=11)
ws1.cell(row=sub_row, column=4, value="$7,503,436").font = Font(bold=True)

ws1.cell(row=sub_row + 2, column=1, value="* El Vite 20x120 no se encuentra online; precios de formato 60x120 como referencia").font = Font(italic=True, size=9)
ws1.cell(row=sub_row + 3, column=1, value="† Precio Bercomat según promo Instagram — verificar vigencia").font = Font(italic=True, size=9)
ws1.cell(row=sub_row + 4, column=1, value="⚡ = Ahorro significativo detectado").font = Font(italic=True, size=9, color="006100")

auto_width(ws1)

# ═══════════════════════════════════════════════════════════════
# SHEET 3: HOJA 2 — SANITARIOS Y VANITORIOS
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Hoja2 - Sanitarios")

ws2.merge_cells('A1:J1')
ws2.cell(row=1, column=1, value="HOJA 2 — PASTINAS, SANITARIOS Y VANITORIOS (netos sin IVA)").font = Font(bold=True, size=14)

headers2 = ["Item", "Producto", "Cant.", "Barugel (neto)", "Blaisten", "Caballito Sanit.", "Bercomat", "Easy", "ML / Otro", "Mejor Precio", "Ahorro vs Barugel"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=2, column=i, value=h)
style_header_row(ws2, 2, len(headers2))

sanit_data = [
    # Pastinas
    ["P.1", "Klaukol Pastina Fluida SKG05 Blanco 5kg", "1 un",
     "$51,629", "$27,678 (s/stk)", "N/D", "N/D", "$19,298", "N/D",
     "$19,298 (Easy)", "-$32,331 ⚡"],
    ["P.2", "Klaukol Pastina Fluida SKG05 Gris Claro 5kg", "1 un",
     "$25,815", "N/D", "N/D", "N/D", "$7,529 (1kg)", "Darsie: $24,197",
     "$24,197 (Darsie)", "-$1,618"],
    ["P.3", "Klaukol Pastina Fluida AP 5kg Perla", "2 un",
     "$27,462", "N/D", "N/D", "N/D", "N/D", "La Clarita: $25,528",
     "$25,528 (La Clarita)", "-$1,934 c/u"],
    # Sanitarios Ferrum
    ["S.1", "Inodoro Largo Ferrum Bari Blanco KLM B", "1 un",
     "$171,094", "$164,455", "N/D", "N/D", "N/D", "Corralón Amigo: $135,285",
     "$135,285 (C. Amigo)", "-$35,809 ⚡"],
    ["S.2", "Depósito Ferrum Bari Blanco de Apoyar", "1 un",
     "$164,721", "$152,719", "N/D", "N/D", "N/D", "Merlino SRL: $163,921",
     "$152,719 (Blaisten)", "-$12,002"],
    ["S.3", "Bidet Ferrum Bari Blanco 3 Agujeros", "1 un",
     "$125,881", "$121,231", "N/D", "N/D", "N/D", "ML: ~$105,000",
     "$105,000 (ML)", "-$20,881 ⚡"],
    ["S.4", "Tapa Asiento Ferrum P/Bari Blanca Madera", "1 un",
     "$58,213", "$141,562*", "N/D", "N/D", "N/D", "Unimax: $47,399",
     "$47,399 (Unimax)", "-$10,814"],
    # Vanitory
    ["V.1", "Schneider Vanitory City Colgar 80cm", "1 un",
     "$559,038", "N/D", "N/D", "N/D", "N/D", "Dimora: $378,226 (s/stk)",
     "$378,226 (Dimora, s/stk)", "—"],
    ["V.2", "Schneider Vanitory Rivo Colg 60cm Blanco", "1 un",
     "$441,636", "N/D", "N/D", "N/D", "N/D", "Merlino SRL: $307,703",
     "$307,703 (Merlino)", "-$133,933 ⚡"],
    # Bañera
    ["B.1", "FV Bianco Bañera Enoxado 150x70", "1 un",
     "$200,741", "N/D", "N/D", "N/D", "N/D", "Ferrum BL 15S: ~$342,872",
     "$200,741 (Barugel)", "$0 (mejor)"],
    ["B.2", "FV Desagüe Lineal 60cm Rejilla Ciega", "1 un",
     "$187,845", "N/D", "N/D", "N/D", "N/D", "Gili y Cía: ~$190,642",
     "$187,845 (Barugel)", "$0 (mejor)"],
]

for r, row_data in enumerate(sanit_data, 3):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        is_best = (c == len(headers2) - 1)
        is_total = (c in [4, len(headers2) - 1, len(headers2)])
        style_data_cell(cell, is_best=is_best, is_total=is_total)
        if c == len(headers2) and '⚡' in str(val):
            cell.fill = BEST_FILL
        if '*' in str(val):
            cell.fill = WARN_FILL

note_row = 3 + len(sanit_data) + 1
ws2.cell(row=note_row, column=1, value="* Tapa en Blaisten de otro material (la de madera de Barugel es más económica)").font = Font(italic=True, size=9)
ws2.cell(row=note_row + 1, column=1, value="Caballito Sanitarios no tiene precios web publicados — cotizar presencialmente (Donato Álvarez 64)").font = Font(italic=True, size=9)
ws2.cell(row=note_row + 2, column=1, value="⚡ = Ahorro significativo detectado").font = Font(italic=True, size=9, color="006100")

auto_width(ws2)

# ═══════════════════════════════════════════════════════════════
# SHEET 4: HOJA 3 — GRIFERÍAS Y ACCESORIOS
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Hoja3 - Griferias")

ws3.merge_cells('A1:I1')
ws3.cell(row=1, column=1, value="HOJA 3 — GRIFERÍAS Y ACCESORIOS FV (netos sin IVA)").font = Font(bold=True, size=14)

headers3 = ["Item", "Producto", "Cant.", "Barugel (neto)", "Blaisten", "Bercomat", "NIMAT", "ML / Otro", "Mejor Precio", "Ahorro vs Barugel"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=2, column=i, value=h)
style_header_row(ws3, 2, len(headers3))

grif_data = [
    ["G.1", "FV Lavatorio Mauna Cromo (JG)", "3 JG",
     "$164,131", "N/D", "N/D", "$185,897", "San. Alvarez: $169,872",
     "$164,131 (Barugel)", "$0 (mejor)"],
    ["G.2", "FV Bidet Mauna Cromo (JG)", "2 JG",
     "$179,038", "N/D", "N/D", "$202,781", "Resta: $205,601",
     "$179,038 (Barugel)", "$0 (mejor)"],
    ["G.3", "FV Ducha c/Transferencia Mauna Cromo (JG)", "2 JG",
     "$187,450", "N/D", "N/D", "$144,690 (ducha)", "N/D",
     "$144,690 (NIMAT)", "-$42,760 ⚡"],
    ["G.4", "Ducha c/Brazo FV Duchamatic Cromo", "2 un",
     "$227,078", "N/D", "N/D", "N/D", "N/D",
     "$227,078 (Barugel)", "$0 (s/ref)"],
    ["A.1", "FV Portarrollos Cromo C/0167", "3 un",
     "$36,242", "N/D", "N/D", "N/D", "N/D",
     "$36,242 (Barugel)", "$0 (s/ref)"],
    ["A.2", "FV Toallero Barra Cromo C/0405", "3 un",
     "$60,404", "N/D", "N/D", "N/D", "N/D",
     "$60,404 (Barugel)", "$0 (s/ref)"],
    ["A.3", "FV Percha Cromo C/0166", "5 un",
     "$22,553", "N/D", "N/D", "N/D", "N/D",
     "$22,553 (Barugel)", "$0 (s/ref)"],
]

for r, row_data in enumerate(grif_data, 3):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        is_best = (c == len(headers3) - 1)
        is_total = (c in [4, len(headers3) - 1, len(headers3)])
        style_data_cell(cell, is_best=is_best, is_total=is_total)
        if c == len(headers3) and '⚡' in str(val):
            cell.fill = BEST_FILL

# Subtotal note
note_r = 3 + len(grif_data) + 1
ws3.cell(row=note_r, column=1, value="NIMAT (Concordia, ER) — precios efectivo/débito. No es CABA pero hace envíos a todo el país.").font = Font(italic=True, size=9)
ws3.cell(row=note_r + 1, column=1, value="Bercomat no tiene precios web visibles — cotizar presencial o por WhatsApp.").font = Font(italic=True, size=9)
ws3.cell(row=note_r + 2, column=1, value="Accesorios FV (portarrollos, toallero, percha) solo encontrados en Barugel — verificar disponibilidad.").font = Font(italic=True, size=9)

auto_width(ws3)

# ═══════════════════════════════════════════════════════════════
# SHEET 5: RESUMEN Y AHORRO POTENCIAL
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Resumen y Ahorros")

ws4.merge_cells('A1:F1')
ws4.cell(row=1, column=1, value="RESUMEN — AHORRO POTENCIAL vs BARUGEL").font = Font(bold=True, size=14)

headers_r = ["Categoría", "Total Barugel (neto)", "Total Mejor Opción (neto)", "Diferencia", "% Ahorro", "Proveedor Recomendado"]
for i, h in enumerate(headers_r, 1):
    ws4.cell(row=2, column=i, value=h)
style_header_row(ws4, 2, len(headers_r))

resumen_data = [
    ["Revestimientos (Klaukol adhesivo)", "$719,612", "$473,200", "-$246,412", "34%", "Bercomat (Klaukol a $18,200)"],
    ["Revest. Portobello Art Martase", "$68,810", "$68,810", "$0", "0%", "Barugel (único proveedor)"],
    ["Tendenza Black Cement", "$91,806", "$90,421", "-$1,385", "1.5%", "Ferrocons / Barugel"],
    ["Eliane Forma Branco", "$485,906", "$485,906", "$0", "0%", "Barugel (mejor precio)"],
    ["Portobello Nord Ris Ripado", "$191,152", "$190,295", "-$857", "0.4%", "Ferrocons"],
    ["Vite Liscio Light Eco Grey", "$154,974", "—", "—", "—", "Formato 20x120 no encontrado"],
    ["Pinto Vivant (Luge + Suage)", "$493,465", "$257,040", "-$236,425", "48% ⚡", "Edificor ($50,400/m²)"],
    ["Pastinas Klaukol (3 items)", "$131,517", "$76,056", "-$55,461", "42% ⚡", "Easy + La Clarita"],
    ["Set Ferrum Bari (4 pzas)", "$519,909", "$440,403", "-$79,506", "15%", "ML + Blaisten + Unimax"],
    ["Schneider Vanitory (City + Rivo)", "$1,000,674", "$685,929", "-$314,745", "31% ⚡", "Merlino SRL (Rivo); Dimora (City s/stk)"],
    ["Bañera + Desagüe FV", "$388,586", "$388,586", "$0", "0%", "Barugel (mejor precio)"],
    ["Grifería FV Mauna (3 JG)", "$530,619", "$530,619", "$0", "0%", "Barugel (lavatorio + bidet)"],
    ["Ducha Mauna c/Transferencia", "$374,900", "$289,380", "-$85,520", "23% ⚡", "NIMAT ($144,690 c/u)"],
    ["Duchamatic c/Brazo", "$454,155", "$454,155", "$0", "0%", "Barugel (s/ref)"],
    ["Accesorios FV (3 items)", "$402,703", "$402,703", "$0", "0%", "Barugel (s/ref)"],
]

total_barugel = 0
total_mejor = 0

for r, row_data in enumerate(resumen_data, 3):
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        is_best = (c == 6)
        style_data_cell(cell, is_best=is_best)
        if c == 4 and '⚡' in str(val):
            cell.fill = BEST_FILL
        # Sum totals from barugel and mejor columns
        if c == 2 and val.startswith('$'):
            try:
                total_barugel += int(val.replace('$', '').replace(',', '').split()[0])
            except:
                pass
        if c == 3 and val.startswith('$'):
            try:
                total_mejor += int(val.replace('$', '').replace(',', '').split()[0])
            except:
                pass

# Total row
total_row = 3 + len(resumen_data) + 1
total_labels = ["TOTAL COMPARABLE", f"${total_barugel:,}", f"${total_mejor:,}", f"-${total_barugel - total_mejor:,}", "—", "—"]
for c, val in enumerate(total_labels, 1):
    cell = ws4.cell(row=total_row, column=c, value=val)
    cell.font = Font(bold=True, size=12)
    cell.fill = GOLD_FILL
    cell.border = THIN_BORDER

# Reference values
ws4.cell(row=total_row + 2, column=1, value="REFERENCIA: Ítem D Arquitectos (Máxima): $8,772,420").font = Font(bold=True)
ws4.cell(row=total_row + 3, column=1, value=f"Barugel Neto Total: $8,092,309").font = Font(bold=True)
ws4.cell(row=total_row + 4, column=1, value=f"Potencial con mejores precios: ~${total_mejor:,} (vs ${total_barugel:,} Barugel)").font = Font(bold=True, color="006100")

# Margin vs limit
margin_barugel = 8772420 - 8092309
potential_total = total_mejor  # this is partial, not all items comparable
ws4.cell(row=total_row + 6, column=1, value=f"MARGEN Barugel vs Límite Máxima: +${margin_barugel:,} libres ✅").font = Font(bold=True, color="006100")
ws4.cell(row=total_row + 7, column=1, value="COMPRAS PENDIENTES DE COTIZAR: Caballito Sanitarios, Bercomat (presencial/WhatsApp)").font = Font(italic=True)
ws4.cell(row=total_row + 8, column=1, value="Cerdisa Deep Sun: NO comparar — precio ya bueno en Barugel, verificar stock con Marcos Badillo").font = Font(italic=True)

auto_width(ws4)

# ═══════════════════════════════════════════════════════════════
# SHEET 6: PULIDO Y PLASTIFICADO — VERMEISTER VELVET
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Pulido y Plastificado")

ws5.merge_cells('A1:G1')
ws5.cell(row=1, column=1, value="ÍTEM E — PULIDO Y PLASTIFICADO DE PISOS (60 m² + escalera)").font = Font(bold=True, size=14)
ws5.cell(row=2, column=1, value="Producto italiano: Vermeister Velvet — Laca poliuretánica monocomp. al agua, efecto soft-touch (gloss 5)").font = Font(italic=True, size=10)

# Section 1: Product specs
ws5.cell(row=4, column=1, value="ESPECIFICACIONES TÉCNICAS — VERMEISTER VELVET").font = Font(bold=True, size=12)
specs = [
    ["Origen", "Italia (Vermeister S.p.A., fundada 1975)"],
    ["Tipo", "Monocomponente al agua, tecnología S-XL"],
    ["Brillo", "Extra mate (gloss 5) — soft-touch"],
    ["Consumo por mano", "80-100 g/m² (≈ 80-100 ml/m²)"],
    ["Rendimiento", "~10-12 m²/L por mano"],
    ["Secado entre manos", "1.5-2 horas"],
    ["Transitable", "12-16 horas"],
    ["Endurecimiento total", "7 días"],
    ["Presentación", "Envase de 5 litros"],
    ["Requiere", "Fondo One (sellador) + Velvet Improver 10% en última mano"],
    ["Certificación", "EC1 (bajo impacto ambiental)"],
]
for r, (k, v) in enumerate(specs, 5):
    ws5.cell(row=r, column=1, value=k).font = Font(bold=True)
    ws5.cell(row=r, column=2, value=v)
    for c in [1, 2]:
        ws5.cell(row=r, column=c).border = THIN_BORDER

# Section 2: Quantity calculation
ws5.cell(row=18, column=1, value="CÁLCULO DE CANTIDADES PARA 60 m² + ESCALERA (~68 m² equivalentes)").font = Font(bold=True, size=12)

calc_headers = ["Concepto", "Fórmula", "Cantidad"]
for i, h in enumerate(calc_headers, 1):
    ws5.cell(row=19, column=i, value=h)
style_header_row(ws5, 19, 3)

calc_data = [
    ["Superficie pisos", "60 m²", "60 m²"],
    ["Escalera (estimado)", "~8 m² equivalentes", "~8 m²"],
    ["Total superficie", "", "~68 m²"],
    ["", "", ""],
    ["Fondo One (sellador) — 1 mano", "68 m² ÷ 10 m²/L", "~6.8 L → 1 × 5L + 1 × 1.5L*"],
    ["Velvet — 1ra mano", "68 m² × 0.09 L/m²", "6.1 L"],
    ["Velvet — 2da mano", "68 m² × 0.09 L/m²", "6.1 L"],
    ["Velvet — 3ra mano (+ Improver)", "68 m² × 0.09 L/m²", "6.1 L"],
    ["Total Velvet (3 manos)", "6.1 L × 3", "18.3 L"],
    ["Velvet Improver (10% última mano)", "6.1 L × 10%", "0.6 L"],
]

for r, row_data in enumerate(calc_data, 20):
    for c, val in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        style_data_cell(cell)
        if r == 29:  # Improver row
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Section 3: Pricing comparison
ws5.cell(row=33, column=1, value="COMPARATIVA DE PRECIOS — SOLO MATERIAL").font = Font(bold=True, size=12)

p_headers = ["Producto", "Cant. necesaria", "Proveedor", "Precio Unit.", "Total", "Stock"]
for i, h in enumerate(p_headers, 1):
    ws5.cell(row=34, column=i, value=h)
style_header_row(ws5, 34, len(p_headers))

pricing_data = [
    ["Velvet 5L", "4 unidades (20L)", "Robledo Pisos", "$343 USD", "$1,372 USD", "S/Stock ❌"],
    ["Velvet 5L", "4 unidades (20L)", "Eternal Parquet (web)", "€145", "€580", "Online"],
    ["Fondo One 5L", "1 unidad (5L)", "Robledo Pisos", "$124 USD", "$124 USD", "S/Stock ❌"],
    ["Fondo One 1L", "1 unidad (est.)", "A consultar", "~$30 USD", "~$30 USD", "—"],
    ["Velvet Improver 0.5L", "1 unidad", "Robledo Pisos", "~$50 USD", "~$50 USD", "—"],
]

for r, row_data in enumerate(pricing_data, 35):
    for c, val in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        style_data_cell(cell)
        if '❌' in str(val):
            cell.fill = WARN_FILL

# Total material
ws5.cell(row=40, column=1, value="TOTAL MATERIAL (Robledo Pisos, USD)").font = Font(bold=True, size=11)
ws5.cell(row=40, column=5, value="~$1,576 USD").font = Font(bold=True, color="C00000")
ws5.cell(row=41, column=1, value="TOTAL MATERIAL (ARS, aprox blue $1.300)").font = Font(bold=True, size=11)
ws5.cell(row=41, column=5, value="~$2,049,000 ARS").font = Font(bold=True, color="C00000")

# Section 4: Service option
ws5.cell(row=43, column=1, value="ALTERNATIVA: SERVICIO COMPLETO (PULIDO + PLASTIFICADO)").font = Font(bold=True, size=12)

serv_headers = ["Tipo de servicio", "Precio m²", "Total 68 m²", "Incluye"]
for i, h in enumerate(serv_headers, 1):
    ws5.cell(row=44, column=i, value=h)
style_header_row(ws5, 44, len(serv_headers))

serv_data = [
    ["Pulido + Plastif. Laca Italiana (ML)", "$23,760/m²", "$1,615,680", "Pulido + 3 manos laca"],
    ["Presupuesto Arquitectos (Ítem E)", "—", "$4,880,000", "Reparación + pulido + plastif. + escalera"],
    ["Diferencia (solo servicio vs arquitectos)", "—", "~$3,264,000", "Arquitectos incluye reparaciones"],
]

for r, row_data in enumerate(serv_data, 45):
    for c, val in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        style_data_cell(cell)
        if r == 46:
            cell.fill = GOLD_FILL

# Section 5: Comparison Vermeister vs Nacional
ws5.cell(row=49, column=1, value="COMPARATIVA VERMEISTER vs ALTERNATIVAS NACIONALES (60 m², 3 manos)").font = Font(bold=True, size=12)

comp_headers = ["Producto", "Origen", "Tipo", "Costo Material", "Toxicidad", "Durabilidad"]
for i, h in enumerate(comp_headers, 1):
    ws5.cell(row=50, column=i, value=h)
style_header_row(ws5, 50, len(comp_headers))

comp_data = [
    ["Vermeister Velvet", "Italia 🇮🇹", "Agua (monocomp.)", "~$2,049,000", "Baja ✅", "Buena (residencial)"],
    ["Vermeister Aqua Play 2", "Italia 🇮🇹", "Agua (bicomponente)", "~$1,080,000", "Baja ✅", "Excelente (alto tránsito)"],
    ["Petrilac Q-501 (Plastilux)", "Argentina 🇦🇷", "Solvente", "~$640,000", "Alta (1 mes) ❌", "Muy buena"],
    ["Kekol K-4035", "Argentina (tec. alemana)", "Agua", "~$370,000", "Baja ✅", "Buena"],
]

for r, row_data in enumerate(comp_data, 51):
    for c, val in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        style_data_cell(cell)
        if r == 51:  # Velvet - highlight
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

ws5.cell(row=56, column=1, value="NOTA: El presupuesto de los arquitectos ($4,880,000) incluye reparaciones + escalera. El servicio de ML ($23,760/m²) es solo pulido+plastificado.").font = Font(italic=True, size=9)
ws5.cell(row=57, column=1, value="El Velvet es el más caro de Vermeister por su acabado soft-touch. Aqua Play 2 es bicomponente, más resistente y casi la mitad de precio.").font = Font(italic=True, size=9, color="C00000")

auto_width(ws5)

# ─── Save ─────────────────────────────────────────────────────
output_path = "/Users/PabloMan/Hermes/projects/active/obra-hortiguera-710/presupuestos/comparativa_proveedores.xlsx"
wb.save(output_path)
print(f"✅ Excel guardado en: {output_path}")
