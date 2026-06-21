#!/usr/bin/env python3
"""
Build Excel comparativo definitivo: Obra Hortiguera 710
Todos los items de Barugel con Top 3 ML + otros proveedores CABA.
TODOS los precios con IVA incluido.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BEST_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_hdr(ws, row, max_col):
    for c in range(1, max_col+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_cell(cell, best=False, warn=False):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    if best: cell.fill = BEST_FILL; cell.font = Font(bold=True)
    if warn: cell.fill = WARN_FILL

def auto_w(ws):
    for col_cells in ws.columns:
        ml = 0
        cl = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                for line in str(cell.value).split('\n'):
                    ml = max(ml, len(line))
        ws.column_dimensions[cl].width = min(max(ml + 3, 10), 50)

# ═══════════════════════════════════════════════
# SHEET: MASTER COMPARATIVA
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = "Comparativa Completa"

ws.merge_cells('A1:H1')
ws.cell(row=1, column=1, value="OBRA HORTIGUERA 710 — COMPARATIVA COMPLETA (precios c/IVA incluido)").font = Font(bold=True, size=14)
ws.cell(row=2, column=1, value="Barugel = referencia. ML#1/2/3 = top 3 MercadoLibre más baratos. Proveedores CABA: Blaisten, Easy, El Amigo, NIMAT, Resta, etc.").font = Font(italic=True, size=9)

hdrs = ["Item", "Producto", "Cant.", "Barugel c/IVA", "ML #1", "ML #2", "ML #3", "Mejor Otro Proveedor", "Mejor Precio Total", "Ahorro vs Barugel"]
for i, h in enumerate(hdrs, 1):
    ws.cell(row=3, column=i, value=h)
style_hdr(ws, 3, len(hdrs))

# Data: [item, producto, cant, barugel_civa, ml1, ml2, ml3, otro, mejor, ahorro]
data = [
    # Hoja 1 - Revestimientos
    ["D.1", "Portobello Art Martase (m²)", "1.44 m²", "$57,819", "—", "—", "—", "No encontrado online", "$57,819", "$0"],
    ["D.2", "Tendenza Black Cement (m²)", "4.95 m²", "$22,442", "$21,177", "$22,307", "$24,786", "Ferrocons: $22,103", "$21,177", "-$1,265"],
    ["D.3", "Eliane Forma Branco 32x60 (m²)", "30.10 m²", "$19,525", "—", "—", "—", "Corralón Mdo: $24,105 / Ranco: $20,727", "$19,525 (Barugel)", "$0"],
    ["D.4", "Portobello Nord Ris Ripado 30x90 (m²)", "4.02 m²", "$57,536", "—", "—", "—", "Ferrocons: $47,337 ⚡", "$47,337", "-$10,199"],
    ["D.5", "Vite Liscio Light Eco Grey (m²)*", "4.32 m²", "$43,406", "$28,872 (120x120)", "$46,128 (60x120)", "$46,641", "Blaisten: $41,940 (60x120)", "$28,872", "— (*formato)"],
    ["D.6", "Pinto Vivant Luge Suave 7x24 (m²)", "2.04 m²", "$117,078", "—", "—", "—", "Edificor: $60,984 ⚡", "$60,984", "-$56,094"],
    ["D.7", "Pinto PBG Vivant Suage 7x24 (m²)", "3.06 m²", "$117,078", "—", "—", "—", "Edificor: $60,984 ⚡", "$60,984", "-$56,094"],
    ["K.1", "Klaukol Adhesivo Porcelanato 25kg", "26 un", "$33,489", "$26,740", "$31,000", "$38,875", "Distabile: $27,712", "$26,740", "-$6,749"],
    # Pastinas
    ["P.1", "Klaukol Pastina Fluida Blanca 5kg", "1 un", "$62,471", "$12,872", "$14,451", "~$21,000", "Easy: $21,390 / La Clarita: $15,975", "$12,872", "-$49,599 ⚡"],
    ["P.2", "Klaukol Pastina Fluida Gris Claro 5kg", "1 un", "$31,236", "~$14,000", "~$16,000", "~$21,000", "Blaisten: $33,490 (Perf)", "~$14,000", "-$17,236"],
    ["P.3", "Klaukol Pastina AP Perla 5kg", "2 un", "$33,229", "$56,030", "—", "—", "La Clarita: $30,888 ⚡", "$30,888", "-$2,341"],
    # Sanitarios Ferrum
    ["S.1", "Inodoro Largo Ferrum Bari Blanco", "1 un", "$207,024", "$165,124", "$185,533", "$209,693", "El Amigo: $163,695 / Blaisten: $198,990", "$163,695", "-$43,329 ⚡"],
    ["S.2", "Depósito Ferrum Bari de Apoyar", "1 un", "$199,312", "$151,348", "$166,635", "$177,100", "El Amigo: $142,194 ⚡ / Blaisten: $184,790", "$142,194", "-$57,118 ⚡"],
    ["S.3", "Bidet Ferrum Bari 3 Agujeros", "1 un", "$152,316", "$103,248", "$146,927", "$150,456", "El Amigo: $120,437 / Blaisten: $146,690", "$103,248", "-$49,068 ⚡"],
    ["S.4", "Tapa Asiento Ferrum Bari Madera", "1 un", "$70,438", "$40,946", "$49,481", "$51,975", "Merlino: $32,912 ⚡ / El Amigo: $62,110", "$32,912", "-$37,526 ⚡"],
    # Vanitory
    ["V.1", "Schneider Vanitory City 80cm Colg", "1 un", "$676,436", "~$599,000", "—", "—", "Dimora: $457,654 (s/stk)", "$457,654", "—"],
    ["V.2", "Schneider Vanitory Rivo 60cm Colg", "1 un", "$534,380", "~$433,000", "—", "—", "Merlino: $307,703 ⚡ / ML: $358,091", "$307,703", "-$226,677 ⚡"],
    # Bañera / Desague
    ["B.1", "FV Bañera Enoxado 150x70", "1 un", "$242,897", "~$363,553", "—", "—", "Barugel es el mejor", "$242,897 (Barugel)", "$0"],
    ["B.2", "FV Desagüe Lineal 60cm", "1 un", "$227,292", "~$212,212", "—", "—", "Gili y Cía: ~$190,642", "$190,642", "-$36,650"],
    # Grifería FV Mauna
    ["G.1", "FV Lavatorio Mauna Cromo (JG)", "3 JG", "$198,599", "$233,485", "—", "—", "NIMAT: $185,897 / Barugel mejor", "$185,897 (NIMAT)", "-$12,702"],
    ["G.2", "FV Bidet Mauna Cromo (JG)", "2 JG", "$216,636", "~$220,577", "—", "—", "NIMAT: $202,781 / Barugel mejor", "$202,781 (NIMAT)", "-$13,855"],
    ["G.3", "FV Ducha Mauna c/Transfer. (JG)", "2 JG", "$226,815", "—", "—", "—", "NIMAT: $144,690 ⚡ (solo ducha)", "$144,690", "-$82,125 ⚡"],
    ["G.4", "FV Duchamatic c/Brazo Cromo", "2 un", "$274,764", "$229,832", "$189,004*", "$154,623*", "Resta: $215,511 (-5% trans)", "$215,511 (Resta)", "-$59,253"],
    # Accesorios FV (corregido: SI hay en ML y Sodimac)
    ["A.1", "FV Portarrollos Cromo C/0167", "3 un", "$43,853", "$17,700 (Arizona ML)", "$19,990", "~$22,000", "Sodimac: $19,599 / FV Arizona", "$17,700", "-$26,153 ⚡"],
    ["A.2", "FV Toallero Barra Cromo C/0405", "3 un", "$73,089", "~$30,000 (Arizona)", "~$35,000", "~$40,000", "Sodimac: ~$35K / combo ML", "$30,000", "-$43,089 ⚡"],
    ["A.3", "FV Percha Cromo C/0166", "5 un", "$27,289", "~$12,000 (Arizona)", "~$15,000", "~$18,000", "Sodimac: ~$12K / combo ML", "$12,000", "-$15,289 ⚡"],
    # Hoja 4 - The Gap (reemplaza Ferrum Bari)
    ["H4.1", "Roca The Gap Inodoro Largo SD", "1 un", "$500,148", "$547,011", "—", "—", "NIMAT: $595,162 (s/depósito)", "$500,148 (Barugel)", "$0"],
    ["H4.2", "Alimentación Inferior The Gap 3/6", "1 un", "$453,198", "~$480,000", "—", "—", "Tucson: $481,422 / NIMAT: $486,953", "$453,198 (Barugel)", "$0"],
    ["H4.3", "Roca The Gap Bidet 3 Agujeros", "1 un", "$169,411", "~$175,000", "~$441,242 (NIMAT)", "~$487,180 (Sagosa)", "Barugel es el mejor", "$169,411 (Barugel)", "$0"],
    ["H4.4", "Tapa The Gap Slim Caída Amortig.", "1 un", "$62,355", "$89,934", "$47,891 (Remat)", "$183,986 (Sagosa)", "Remateriales: $39,270 (-18%)", "$39,270", "-$23,085 ⚡"],
    # Espejos
    ["E.1", "Reflejar Rect. 100×80 LED Touch", "1 un", "$297,814", "$259,100", "—", "—", "Marash: $276,800 / Reflejar web: $318,960", "$259,100", "-$38,714"],
    ["E.2", "Schneider Espejo LED", "1 un", "$282,385", "~$200-260K", "—", "—", "Schneider web: varios modelos", "~$200,000", "-$82,385"],
    ["E.3", "Gemini 60×70 Borde Acrílico LED", "1 un", "$611,290", "$259,100", "$263,659 (Bonomi)", "$299,750", "ML: $259,100 ⚡", "$259,100", "-$352,190 ⚡"],
]

for r, row in enumerate(data, 4):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        is_best = (c == 9)
        style_cell(cell, best=is_best, warn=('—' in str(val) and c >= 5))
        if c == 10 and '⚡' in str(val):
            cell.fill = BEST_FILL
            cell.font = Font(bold=True, color="006100")
        # Color alterno filas
        if r % 2 == 0:
            if not cell.fill or cell.fill == PatternFill():
                cell.fill = GRAY_FILL

auto_w(ws)

# ═══════════════════════════════════════════════
# SHEET: RESUMEN AHORRO
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("Resumen Ahorro")

ws2.merge_cells('A1:E1')
ws2.cell(row=1, column=1, value="RESUMEN DE AHORRO vs BARUGEL (comprando al mejor precio por producto)").font = Font(bold=True, size=13)
ws2.merge_cells('A2:E2')
ws2.cell(row=2, column=1, value="Ahorro potencial total: armando la compra entre varios proveedores (no todo en Barugel)").font = Font(italic=True, size=10)

rhdrs = ["Categoría", "Costo Barugel c/IVA", "Costo Mejor Precio", "Ahorro", "%"]
for i, h in enumerate(rhdrs, 1):
    ws2.cell(row=3, column=i, value=h)
style_hdr(ws2, 3, len(rhdrs))

resumen = [
    ["Revestimientos (sin Cerdisa)", "$2,700,156", "~$1,898,000", "~$802,000", "30%"],
    ["Pastinas Klaukol", "$126,937", "~$58,000", "~$69,000", "54%"],
    ["Set Ferrum Bari (4 pzas)", "$629,090", "~$442,049", "~$187,000", "30%"],
    ["Vanitory Schneider (2 un)", "$1,210,816", "~$765,357", "~$445,000", "37%"],
    ["Bañera + Desague FV", "$470,189", "~$433,539", "~$37,000", "8%"],
    ["Grifería FV Mauna (3 JG)", "$642,050", "~$533,368", "~$109,000", "17%"],
    ["Duchamatic (2 un)", "$549,528", "~$431,022", "~$119,000", "22%"],
    ["Accesorios FV (3 items)", "$144,231", "~$60,000", "~$84,000", "58% ⚡"],
    ["Set Roca The Gap (4 pzas)", "$1,185,112", "~$1,162,000", "~$23,000", "2%"],
    ["Espejos (3 un)", "$1,191,489", "~$718,000", "~$473,000", "40% ⚡"],
]

total_bar = 0
total_mejor = 0
for r, row in enumerate(resumen, 4):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        style_cell(cell)
        if c == 4 and not val.startswith("$0"):
            cell.fill = BEST_FILL
        if c == 5:
            try:
                pct = int(val.replace('%', ''))
                if pct >= 20:
                    cell.fill = BEST_FILL
                    cell.font = Font(bold=True, color="006100")
            except:
                pass
    # Sum totals
    try:
        total_bar += int(row[1].replace('$', '').replace(',', '').split('~')[1] if '~' in row[1] else row[1].replace('$', '').replace(',', ''))
    except:
        pass
    try:
        total_mejor += int(row[2].replace('$', '').replace(',', '').split('~')[1] if '~' in row[2] else row[2].replace('$', '').replace(',', ''))
    except:
        pass

tr = 4 + len(resumen) + 1
for c, val in enumerate(["TOTAL", f"${total_bar:,}", f"${total_mejor:,}", f"-${total_bar - total_mejor:,}", f"{int((1-total_mejor/total_bar)*100)}%"], 1):
    cell = ws2.cell(row=tr, column=c, value=val)
    cell.font = Font(bold=True, size=12)
    cell.fill = GOLD_FILL
    cell.border = THIN_BORDER

ws2.cell(row=tr+2, column=1, value="Ítem D Arquitectos (límite): $8,772,420   |   Barugel neto total: $8,092,309   |   Margen libre: $680,111 ✅").font = Font(bold=True, color="006100")
ws2.cell(row=tr+3, column=1, value="Cerdisa Deep Sun: NO comparar — precio ya bueno en Barugel, verificar stock").font = Font(italic=True, size=9)
ws2.cell(row=tr+4, column=1, value="Pendientes: Caballito Sanitarios (presencial), Bercomat (verificar promo Klaukol), negociar flete Marcos Badillo").font = Font(italic=True, size=9)

auto_w(ws2)

# Save
out = "/Users/PabloMan/Hermes/projects/active/obra-hortiguera-710/presupuestos/comparativa_completa.xlsx"
wb.save(out)
print(f"✅ Excel guardado en: {out}")
